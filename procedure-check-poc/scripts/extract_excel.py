#!/usr/bin/env python3
"""Excel 手順書を AI 判定用の Markdown に整形する。

依存: openpyxl のみ（社内環境でも入手しやすい最小構成）
使い方:
    python3 extract_excel.py 手順書.xlsx -o procedure.md
"""
import argparse
import sys

from openpyxl import load_workbook


def cell_text(value) -> str:
    if value is None:
        return ""
    text = str(value)
    # Markdown テーブルを壊さないようにセル内改行・パイプを退避
    return text.replace("|", "\\|").replace("\r\n", "<br>").replace("\n", "<br>").strip()


def expand_merged(ws):
    """結合セルの左上値を全域に展開した2次元リストを返す。"""
    grid = [[cell_text(c.value) for c in row] for row in ws.iter_rows()]
    for rng in ws.merged_cells.ranges:
        top = grid[rng.min_row - 1][rng.min_col - 1]
        for r in range(rng.min_row - 1, rng.max_row):
            for c in range(rng.min_col - 1, rng.max_col):
                grid[r][c] = top
    return grid


def trim_grid(grid):
    """末尾の空行・空列を落とす。"""
    while grid and all(v == "" for v in grid[-1]):
        grid.pop()
    if not grid:
        return grid
    width = max(len(r) for r in grid)
    grid = [r + [""] * (width - len(r)) for r in grid]
    while width > 0 and all(r[width - 1] == "" for r in grid):
        width -= 1
    return [r[:width] for r in grid]


def sheet_to_markdown(ws) -> str:
    grid = trim_grid(expand_merged(ws))
    if not grid:
        return "（空シート）\n"
    lines = []
    header, *body = grid
    lines.append("| " + " | ".join(header) + " |")
    lines.append("|" + "|".join([" --- "] * len(header)) + "|")
    for row in body:
        lines.append("| " + " | ".join(row) + " |")
    return "\n".join(lines) + "\n"


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("xlsx", help="手順書 Excel ファイル")
    parser.add_argument("-o", "--output", help="出力先（省略時は標準出力）")
    args = parser.parse_args()

    wb = load_workbook(args.xlsx, data_only=True)
    parts = [f"# 手順書: {args.xlsx}\n"]
    for ws in wb.worksheets:
        if ws.sheet_state != "visible":
            continue
        parts.append(f"\n## シート: {ws.title}\n")
        parts.append(sheet_to_markdown(ws))
    text = "\n".join(parts)

    if args.output:
        with open(args.output, "w", encoding="utf-8") as f:
            f.write(text)
    else:
        sys.stdout.write(text)
    return 0


if __name__ == "__main__":
    sys.exit(main())
