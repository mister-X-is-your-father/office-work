# -*- coding: utf-8 -*-
"""会議ファシリテーション・テンプレ（3タブ・最強版）
『世界で一番やさしい会議の教科書』(榊巻亮) の手法を、1枚の会議シートを上から
埋めるだけで実践できる形に。論点リスト＋カバレッジ＋MECE観点チェックで「抜け漏れ」を
構造的に防ぐ。決定/論点の自動集約・時間予実・会議コスト・ファシリ点数を数式で自動化。"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import FormulaRule, CellIsRule
from openpyxl.utils import get_column_letter

NAVY="1F3A5F"; SLATE="334155"; TEAL="0F766E"; BLUE="1D4ED8"; PURPLE="6D28D9"
GREEN="15803D"; AMBER="B45309"; ROSE="BE123C"; RED="B91C1C"
C_OPINION="DBEAFE"; C_ISSUE="FEF3C7"; C_DECISION="DCFCE7"
C_DONE="DCFCE7"; C_OVERDUE="FEE2E2"; C_INPUT="FFFEF7"; C_DIV="F1F5F9"
C_SAY="EEF2FF"; C_EX="94A3B8"; WHITE="FFFFFF"; INK="0F172A"

thin=Side(style="thin", color="CBD5E1")
BORDER=Border(left=thin,right=thin,top=thin,bottom=thin)
def fill(h): return PatternFill("solid", fgColor=h)

wb=Workbook()
S_MTG="▶ 会議"; S_LEDGER="🗂 課題・タスク台帳"; S_COVER="🧭 網羅チェック"; S_GUIDE="📖 ガイド"

def newsheet(t,c):
    ws=wb.create_sheet(t); ws.sheet_properties.tabColor=c; ws.sheet_view.showGridLines=False; return ws
def title_row(ws,text,span,color):
    ws.merge_cells(f"A1:{get_column_letter(span)}1"); c=ws["A1"]; c.value=text
    c.font=Font(bold=True,size=15,color=WHITE); c.fill=fill(color)
    c.alignment=Alignment(vertical="center",horizontal="left",indent=1); ws.row_dimensions[1].height=34
def sub_row(ws,text,span):
    ws.merge_cells(f"A2:{get_column_letter(span)}2"); c=ws["A2"]; c.value=text
    c.font=Font(size=10,italic=True,color=SLATE)
    c.alignment=Alignment(vertical="center",horizontal="left",indent=1,wrap_text=True); ws.row_dimensions[2].height=28
def section(ws,row,text,span,color):
    ws.merge_cells(f"A{row}:{get_column_letter(span)}{row}"); c=ws[f"A{row}"]; c.value=text
    c.font=Font(bold=True,size=11,color=WHITE); c.fill=fill(color)
    c.alignment=Alignment(vertical="center",horizontal="left",indent=1); ws.row_dimensions[row].height=24
def cell(ws,row,col,val="",bold=False,italic=False,color=INK,fillhex=None,align="left",wrap=True,border=True,size=10):
    c=ws.cell(row=row,column=col,value=val)
    c.font=Font(size=size,bold=bold,italic=italic,color=color)
    c.alignment=Alignment(vertical="top",horizontal=align,wrap_text=wrap,indent=(1 if align=="left" else 0))
    if fillhex: c.fill=fill(fillhex)
    if border: c.border=BORDER
    return c
def label(ws,row,col,text): return cell(ws,row,col,text,bold=True,fillhex=C_DIV)
def inp(ws,row,col,ph=""): return cell(ws,row,col,ph,italic=bool(ph),color=(C_EX if ph else INK),fillhex=C_INPUT,align="left")
def say(ws,row,col,text): return cell(ws,row,col,text,italic=True,color=SLATE,fillhex=C_SAY)
def widths(ws,m):
    for k,v in m.items(): ws.column_dimensions[k].width=v
def hdr(ws,row,cells,color):
    for rng,text in cells:
        if ":" in rng:
            ws.merge_cells(rng); a=rng.split(":")[0]; r=ws[a].row; col=ws[a].column
        else:
            r=row; col=ws[rng].column
        c=ws.cell(row=r,column=col,value=text)
        c.font=Font(bold=True,size=10,color=WHITE); c.fill=fill(color); c.border=BORDER
        c.alignment=Alignment(vertical="center",horizontal="center",wrap_text=True)
    ws.row_dimensions[row].height=22

# =========================================================
# ▶ 会議（counter-based dynamic layout）
# =========================================================
ws=newsheet(S_MTG, NAVY)
widths(ws,{"A":11,"B":13,"C":22,"D":16,"E":11,"F":12})
title_row(ws,"▶ 会議シート — このまま上から埋める＝会議が回る",6,NAVY)
sub_row(ws,"終わったら右クリック→『シートを複製』で次回へ。③のToDoは🗂台帳へ転記。論点カバレッジ・決定・時間・コストは自動連携。",6)
R=3
label(ws,R,1,"会議名"); ws.merge_cells(f"B{R}:C{R}"); inp(ws,R,2,"（記入）")
label(ws,R,4,"日時"); ws.merge_cells(f"E{R}:F{R}"); inp(ws,R,5,""); R+=1
label(ws,R,1,"ファシリ"); ws.merge_cells(f"B{R}:C{R}"); inp(ws,R,2,"")
label(ws,R,4,"書記"); ws.merge_cells(f"E{R}:F{R}"); inp(ws,R,5,""); R+=1
# コスト dashboard（式は合計確定後に入れる）
label(ws,R,1,"人数(名)"); inp(ws,R,2,""); PEOPLE=f"B{R}"
label(ws,R,3,"時給(円)"); cell(ws,R,4,3000,align="center",fillhex=C_INPUT); WAGE=f"D{R}"
label(ws,R,5,"想定コスト"); ccost=cell(ws,R,6,"",align="center",bold=True,color=AMBER,fillhex="FFF7ED"); ccost.number_format='"¥"#,##0'; R+=1
# ① 準備
section(ws,R,"① 準備 ─ 終了条件・型・アジェンダ・論点出し",6,BLUE); R+=1
label(ws,R,1,"終了条件"); ws.merge_cells(f"B{R}:F{R}"); inp(ws,R,2,"（例: 来期の販促3案から1案に絞り、担当と期限が決まっている）"); PURPOSE=f"B{R}"; R+=1
label(ws,R,1,"言うことば"); ws.merge_cells(f"B{R}:F{R}"); say(ws,R,2,"冒頭で音読 →『今日のゴールは“〜”です。これでいいですか?』"); R+=1
label(ws,R,1,"会議の型"); inp(ws,R,2,"▼選択"); ws.merge_cells(f"C{R}:F{R}"); cell(ws,R,3,"型ごとの進め方は📖ガイド参照",italic=True,color=C_EX)
dv=DataValidation(type="list",formula1='"A 報告,B 情報収集,C 承認,D 方針検討,E 課題解決"',allow_blank=True); ws.add_data_validation(dv); dv.add(f"B{R}"); R+=1
hdr(ws,R,[(f"A{R}","#"),(f"B{R}:D{R}","議題（＋ゴール）"),(f"E{R}","予定(分)"),(f"F{R}","実績(分)")],BLUE); R+=1
AG_S=R; ex_ag=[("（例）前回ToDoの振り返り：完了/未完を確認",5)]
for i in range(5):
    cell(ws,R,1,i+1,align="center"); ws.merge_cells(f"B{R}:D{R}")
    if i<len(ex_ag): cell(ws,R,2,ex_ag[i][0],italic=True,color=C_EX,fillhex=C_INPUT); cell(ws,R,5,ex_ag[i][1],align="center",italic=True,color=C_EX,fillhex=C_INPUT); cell(ws,R,6,"",align="center",fillhex=C_INPUT)
    else: inp(ws,R,2,""); cell(ws,R,5,"",align="center",fillhex=C_INPUT); cell(ws,R,6,"",align="center",fillhex=C_INPUT)
    R+=1
AG_E=R-1
ws.merge_cells(f"A{R}:D{R}"); cell(ws,R,1,"合計",bold=True,align="right",fillhex=C_DIV)
cell(ws,R,5,f"=SUM(E{AG_S}:E{AG_E})",align="center",bold=True,fillhex=C_DIV); cell(ws,R,6,f"=SUM(F{AG_S}:F{AG_E})",align="center",bold=True,fillhex=C_DIV)
PLAN=f"E{R}"; ACT=f"F{R}"; R+=1
ccost.value=f'=IFERROR({PEOPLE}*({PLAN}/60)*{WAGE},"")'
# ■ 論点リスト（抜け漏れ対策の核）
section(ws,R,"■ 論点リスト ─ 先に『決めること』を全部出す → カバレッジを追う（抜け漏れ対策）",6,BLUE); R+=1
label(ws,R,1,"言うことば"); ws.merge_cells(f"B{R}:F{R}"); say(ws,R,2,"冒頭で『今日決めることは①②③…これで全部ですか?』と合意。各論点を“結論”まで持っていく。"); R+=1
hdr(ws,R,[(f"A{R}","#"),(f"B{R}:C{R}","論点（決めること/議論すること）"),(f"D{R}","状態"),(f"E{R}:F{R}","結論メモ")],BLUE); R+=1
IS_S=R; ex_is=[("販促予算の上限は?","結論","50万円を上限"),("どのチャネルを使う?","議論中",""),("担当と期限は?","未","")]
for i in range(6):
    cell(ws,R,1,i+1,align="center"); ws.merge_cells(f"B{R}:C{R}"); ws.merge_cells(f"E{R}:F{R}")
    if i<len(ex_is):
        t,st,memo=ex_is[i]; cell(ws,R,2,t,italic=True,color=C_EX,fillhex=C_INPUT); cell(ws,R,4,st,align="center",italic=True,color=C_EX,fillhex=C_INPUT); cell(ws,R,5,memo,italic=True,color=C_EX,fillhex=C_INPUT)
    else: inp(ws,R,2,""); cell(ws,R,4,"",align="center",fillhex=C_INPUT); inp(ws,R,5,"")
    R+=1
IS_E=R-1; ISD=f"D{IS_S}:D{IS_E}"; ISB=f"B{IS_S}:B{IS_E}"
dv=DataValidation(type="list",formula1='"未,議論中,結論"',allow_blank=True); ws.add_data_validation(dv); dv.add(ISD)
ws.conditional_formatting.add(f"A{IS_S}:F{IS_E}",FormulaRule(formula=[f'$D{IS_S}="結論"'],fill=fill(C_DECISION)))
ws.conditional_formatting.add(f"A{IS_S}:F{IS_E}",FormulaRule(formula=[f'$D{IS_S}="議論中"'],fill=fill(C_ISSUE)))
label(ws,R,1,"カバレッジ"); ws.merge_cells(f"B{R}:F{R}")
cell(ws,R,2,f'=COUNTIF({ISD},"結論")&" / "&COUNTA({ISB})&" 論点に結論　｜　未対応 "&(COUNTA({ISB})-COUNTIF({ISD},"結論"))&" 件"',bold=True,fillhex=C_INPUT)
ws.conditional_formatting.add(f"B{R}",FormulaRule(formula=[f'(COUNTA({ISB})-COUNTIF({ISD},"結論"))=0'],fill=fill(C_DONE),font=Font(bold=True,color=GREEN)))
ws.conditional_formatting.add(f"B{R}",FormulaRule(formula=[f'(COUNTA({ISB})-COUNTIF({ISD},"結論"))>0'],fill=fill(C_OVERDUE),font=Font(bold=True,color=RED))); R+=1
# ■ 抜け漏れ観点チェック（MECEレンズ）
section(ws,R,"■ 抜け漏れ観点チェック ─ 発散の最後に『他にない?』（MECEレンズ）",6,BLUE); R+=1
label(ws,R,1,"観点点数"); ws.merge_cells(f"B{R}:F{R}"); LSCORE=R; R+=1
LENS_S=R; lenses=["関係者の視点（誰の立場が抜けてない?）","選択肢（他の案は? 出し切った?）","リスク・反対意見（マイナス面は?）","時間軸（短期/長期 両方見た?）","ヒト・モノ・カネ・情報（資源の観点）","前提・制約（条件は確認した?）"]
for i,l in enumerate(lenses):
    cell(ws,R,1,i+1,align="center",fillhex=C_DIV); cell(ws,R,2,"",align="center",fillhex=C_INPUT); ws.merge_cells(f"C{R}:F{R}"); cell(ws,R,3,l); R+=1
LENS_E=R-1
dv=DataValidation(type="list",formula1='"✓"',allow_blank=True); ws.add_data_validation(dv); dv.add(f"B{LENS_S}:B{LENS_E}")
ws.conditional_formatting.add(f"A{LENS_S}:F{LENS_E}",FormulaRule(formula=[f'$B{LENS_S}="✓"'],fill=fill(C_DONE)))
cell(ws,LSCORE,2,f'=COUNTIF(B{LENS_S}:B{LENS_E},"✓")&" / 6 観点 確認済"',bold=True,color=TEAL,fillhex=C_INPUT)
# ② 進行
section(ws,R,"② 進行 ─ スクライブ｜意見=青/論点=黄/決定=緑（自動色分け）",6,PURPLE); R+=1
label(ws,R,1,"発散/収束"); ws.merge_cells(f"B{R}:F{R}"); say(ws,R,2,"今は【発散：出す】/【収束：決める】← 宣言してから。迷子は『事象→問題→原因→施策→効果』の今どこ?"); R+=1
hdr(ws,R,[(f"A{R}","時刻"),(f"B{R}","種別"),(f"C{R}:E{R}","内容"),(f"F{R}","発言者")],PURPLE); R+=1
SC_S=R; ex_sc=[("10:05","論点","販促予算の上限は?","田中"),("10:07","意見","A案はSNS中心で低コスト","鈴川"),("10:12","決定","予算は50万円を上限とする","合意")]
for i in range(20):
    ws.merge_cells(f"C{R}:E{R}")
    if i<len(ex_sc):
        t,k,c,w=ex_sc[i]; cell(ws,R,1,t,align="center",italic=True,color=C_EX); cell(ws,R,2,k,align="center",italic=True,color=C_EX); cell(ws,R,3,c,italic=True,color=C_EX); cell(ws,R,6,w,align="center",italic=True,color=C_EX)
    else:
        cell(ws,R,1,"",align="center",fillhex=C_INPUT); cell(ws,R,2,"",align="center",fillhex=C_INPUT); cell(ws,R,3,"",fillhex=C_INPUT); cell(ws,R,6,"",align="center",fillhex=C_INPUT)
    R+=1
SC_E=R-1
dv=DataValidation(type="list",formula1='"意見,論点,決定"',allow_blank=True); ws.add_data_validation(dv); dv.add(f"B{SC_S}:B{SC_E}")
ws.conditional_formatting.add(f"A{SC_S}:F{SC_E}",FormulaRule(formula=[f'$B{SC_S}="決定"'],fill=fill(C_DECISION),font=Font(bold=True)))
ws.conditional_formatting.add(f"A{SC_S}:F{SC_E}",FormulaRule(formula=[f'$B{SC_S}="論点"'],fill=fill(C_ISSUE)))
ws.conditional_formatting.add(f"A{SC_S}:F{SC_E}",FormulaRule(formula=[f'$B{SC_S}="意見"'],fill=fill(C_OPINION)))
# 🅿 駐車場
section(ws,R,"🅿 駐車場 ─ 脱線ネタを退避（あとで拾う）",6,SLATE); R+=1
for i in range(3):
    cell(ws,R,1,"・",align="center",fillhex=C_INPUT); ws.merge_cells(f"B{R}:F{R}"); inp(ws,R,2,""); R+=1
# ③ 確認
section(ws,R,"③ 確認 ─ 終わる前の5分（決定・未対応論点が自動集約）",6,GREEN); R+=1
label(ws,R,1,"言うことば"); ws.merge_cells(f"B{R}:F{R}"); say(ws,R,2,"『決まったことは○○ですね』『やるべきことは、誰が・いつまでに・何を?』『論点は全部つぶれましたか?』"); R+=1
label(ws,R,1,"論点カバレッジ"); ws.merge_cells(f"B{R}:F{R}")
cell(ws,R,2,f'="未対応の論点 "&(COUNTA({ISB})-COUNTIF({ISD},"結論"))&" 件　（0なら抜け漏れOK）"',bold=True,fillhex=C_INPUT)
ws.conditional_formatting.add(f"B{R}",FormulaRule(formula=[f'(COUNTA({ISB})-COUNTIF({ISD},"結論"))=0'],fill=fill(C_DONE),font=Font(bold=True,color=GREEN)))
ws.conditional_formatting.add(f"B{R}",FormulaRule(formula=[f'(COUNTA({ISB})-COUNTIF({ISD},"結論"))>0'],fill=fill(C_OVERDUE),font=Font(bold=True,color=RED))); R+=1
label(ws,R,1,"未対応(自動)"); ws.merge_cells(f"B{R}:F{R}")
cell(ws,R,2,f'=IFERROR("? "&TEXTJOIN(CHAR(10)&"? ",TRUE,FILTER({ISB},{ISD}<>"結論")),"（論点リストが全部『結論』＝抜け漏れなし）")',fillhex=C_ISSUE); ws.row_dimensions[R].height=52; R+=1
label(ws,R,1,"決定(自動)"); ws.merge_cells(f"B{R}:F{R}")
cell(ws,R,2,f'=IFERROR("・ "&TEXTJOIN(CHAR(10)&"・ ",TRUE,FILTER(C{SC_S}:C{SC_E},B{SC_S}:B{SC_E}="決定")),"（②で種別を『決定』にすると自動表示）")',fillhex=C_DECISION); ws.row_dimensions[R].height=60; R+=1
hdr(ws,R,[(f"A{R}","#"),(f"B{R}","担当"),(f"C{R}:E{R}","やること（何を）"),(f"F{R}","期限")],GREEN); R+=1
for i in range(4):
    cell(ws,R,1,i+1,align="center"); inp(ws,R,2,""); ws.merge_cells(f"C{R}:E{R}"); inp(ws,R,3,""); cell(ws,R,6,"",align="center",fillhex=C_INPUT); ws[f"F{R}"].number_format="yyyy-mm-dd"; R+=1
label(ws,R,1,"終了条件"); ws.merge_cells(f"B{R}:F{R}"); cell(ws,R,2,f"={PURPOSE}",italic=True,color=SLATE,fillhex=C_INPUT); R+=1
label(ws,R,1,"達成度"); inp(ws,R,2,"▼"); ws.merge_cells(f"C{R}:F{R}"); inp(ws,R,3,"未達なら→ 次アクションを記入")
dv=DataValidation(type="list",formula1='"達成,一部,未達"',allow_blank=True); ws.add_data_validation(dv); dv.add(f"B{R}")
ws.conditional_formatting.add(f"B{R}",CellIsRule(operator="equal",formula=['"達成"'],fill=fill(C_DECISION),font=Font(bold=True)))
ws.conditional_formatting.add(f"B{R}",CellIsRule(operator="equal",formula=['"一部"'],fill=fill(C_ISSUE)))
ws.conditional_formatting.add(f"B{R}",CellIsRule(operator="equal",formula=['"未達"'],fill=fill(C_OVERDUE))); R+=1
# ④ 振り返り
section(ws,R,"④ 振り返り ─ 次回へ（時間の予実は自動）",6,ROSE); R+=1
label(ws,R,1,"予定(分)"); cell(ws,R,2,f"={PLAN}",align="center",fillhex=C_INPUT)
label(ws,R,3,"実績(分)"); cell(ws,R,4,f"={ACT}",align="center",fillhex=C_INPUT)
label(ws,R,5,"差"); cell(ws,R,6,f"=D{R}-B{R}",align="center",bold=True); R+=1
hdr(ws,R,[(f"A{R}:B{R}","K 続ける"),(f"C{R}:D{R}","P 課題"),(f"E{R}:F{R}","T 次に試す")],ROSE); R+=1
for i in range(3):
    ws.merge_cells(f"A{R}:B{R}"); cell(ws,R,1,"",fillhex=C_INPUT); ws.merge_cells(f"C{R}:D{R}"); cell(ws,R,3,"",fillhex=C_INPUT); ws.merge_cells(f"E{R}:F{R}"); cell(ws,R,5,"",fillhex=C_INPUT); R+=1
label(ws,R,1,"次回"); ws.merge_cells(f"B{R}:F{R}"); inp(ws,R,2,"日時・主な議題"); R+=1
# ✅ ファシリ・チェック（8つの基本動作）
section(ws,R,"✅ 今日のファシリ・チェック ─ 8つの基本動作（✓した数が今日の点数）",6,TEAL); R+=1
label(ws,R,1,"今日の点数"); ws.merge_cells(f"B{R}:F{R}"); CSCORE=R; R+=1
ACT_S=R; acts=["決まったこと・やるべきことを確認した","終了条件を最初に確認・宣言した","時間配分を宣言した","議論を可視化(スクライブ)した","発散/収束を切り替え宣言した","主張を引き出した（黙ってる人に振った）","対話を促した（一方通行にしない）","振り返りをした"]
for i,a in enumerate(acts):
    cell(ws,R,1,i+1,align="center",fillhex=C_DIV); cell(ws,R,2,"",align="center",fillhex=C_INPUT); ws.merge_cells(f"C{R}:F{R}"); cell(ws,R,3,a); R+=1
ACT_E=R-1
dv=DataValidation(type="list",formula1='"✓"',allow_blank=True); ws.add_data_validation(dv); dv.add(f"B{ACT_S}:B{ACT_E}")
ws.conditional_formatting.add(f"A{ACT_S}:F{ACT_E}",FormulaRule(formula=[f'$B{ACT_S}="✓"'],fill=fill(C_DONE)))
cell(ws,CSCORE,2,f'=COUNTIF(B{ACT_S}:B{ACT_E},"✓")&" / 8"',bold=True,color=TEAL,fillhex=C_INPUT)
ws.freeze_panes="A3"

# =========================================================
# 🗂 課題・タスク台帳（マスター）＋ 棚卸しダッシュボード
# =========================================================
ws=newsheet(S_LEDGER, AMBER)
widths(ws,{"A":5,"B":13,"C":10,"D":36,"E":11,"F":13,"G":9,"H":11,"I":12,"J":6,"K":11,"L":26})
title_row(ws,"🗂 課題・タスク台帳（マスター）— やるべき/議論すべきを全部ここに貯める",12,AMBER)
sub_row(ws,"会議で出た決定・タスク・論点は必ずここへ＝落とさない唯一の箱。上部の棚卸し（要対応）は自動表示。",12)
DS=6; DE=205; LE=DS+39
cards=[("未完", f'=COUNTIFS(I{DS}:I{DE},"<>完了",D{DS}:D{DE},"<>")&" 件"'),
       ("期限超過", f'=COUNTIFS(H{DS}:H{DE},"<"&TODAY(),H{DS}:H{DE},"<>",I{DS}:I{DE},"<>完了")&" 件"'),
       ("停滞(>14日)", f'=SUMPRODUCT((K{DS}:K{DE}<>"")*(I{DS}:I{DE}<>"完了")*((TODAY()-K{DS}:K{DE})>14))&" 件"'),
       ("担当なし", f'=COUNTIFS(G{DS}:G{DE},"",D{DS}:D{DE},"<>",I{DS}:I{DE},"<>完了")&" 件"'),
       ("議論待ち", f'=COUNTIF(I{DS}:I{DE},"議論待ち")&" 件"'),
       ("完了率", f'=IFERROR(COUNTIF(I{DS}:I{DE},"完了")/COUNTA(A{DS}:A{DE}),"")')]
col=1
for lab,frm in cards:
    L1=get_column_letter(col); L2=get_column_letter(col+1)
    ws.merge_cells(f"{L1}3:{L2}3"); cell(ws,3,col,lab,bold=True,align="center",fillhex=C_DIV)
    ws.merge_cells(f"{L1}4:{L2}4"); cell(ws,4,col,frm,bold=True,align="center",fillhex=C_INPUT,color=AMBER)
    col+=2
ws["K4"].number_format="0%"
hdr(ws,5,[("A5","ID"),("B5","領域"),("C5","種別"),("D5","内容"),("E5","起票日"),("F5","起票元(会議)"),("G5","担当"),("H5","期限"),("I5","状態"),("J5","優先"),("K5","最終更新"),("L5","次アクション")],AMBER)
ex=[(1,"スケジュール","タスク","A案の見積を取得","2026-05-30","販促キックオフ","田中","2026-06-06","進行中","高","2026-05-30","見積依頼済")]
for i in range(40):
    r=DS+i
    if i<len(ex):
        for c0,v in enumerate(ex[i]):
            al="center" if c0 in (0,4,7,8,9,10) else "left"; cell(ws,r,c0+1,v,italic=True,color=C_EX,align=al)
    else:
        cell(ws,r,1,i+1,align="center")
        for c0 in range(2,13):
            al="center" if c0 in (3,5,8,9,10,11) else "left"; cell(ws,r,c0,"",align=al,fillhex=C_INPUT)
    ws[f"E{r}"].number_format="yyyy-mm-dd"; ws[f"H{r}"].number_format="yyyy-mm-dd"; ws[f"K{r}"].number_format="yyyy-mm-dd"
for rng,opts in [(f"B{DS}:B{LE}",'"目的・スコープ,成果物・要件,スケジュール,体制・役割,予算,リスク・課題,品質,関係者・コミュ,依存・前提,意思決定・承認,運用・保守,法務・セキュリティ,その他"'),
                 (f"C{DS}:C{LE}",'"タスク,論点,リスク,依存,決定"'),
                 (f"I{DS}:I{LE}",'"未着手,進行中,完了,保留,議論待ち"'),
                 (f"J{DS}:J{LE}",'"高,中,低"')]:
    dv=DataValidation(type="list",formula1=opts,allow_blank=True); ws.add_data_validation(dv); dv.add(rng)
allrng=f"A{DS}:L{LE}"
ws.conditional_formatting.add(allrng,FormulaRule(formula=[f'$I{DS}="完了"'],fill=fill(C_DONE)))
ws.conditional_formatting.add(allrng,FormulaRule(formula=[f'$I{DS}="議論待ち"'],fill=fill(C_OPINION)))
ws.conditional_formatting.add(f"H{DS}:H{LE}",FormulaRule(formula=[f'AND($H{DS}<>"",$I{DS}<>"完了",$H{DS}<TODAY())'],fill=fill(C_OVERDUE),font=Font(bold=True,color=RED)))
ws.conditional_formatting.add(f"K{DS}:K{LE}",FormulaRule(formula=[f'AND($K{DS}<>"",$I{DS}<>"完了",(TODAY()-$K{DS})>14)'],fill=fill(C_ISSUE)))
ws.conditional_formatting.add(f"G{DS}:G{LE}",FormulaRule(formula=[f'AND($G{DS}="",$D{DS}<>"",$I{DS}<>"完了")'],fill=fill(C_OVERDUE)))
ws.freeze_panes="A6"

# =========================================================
# 🧭 網羅チェック（領域マップ＝やるべき/議論すべきの基準）
# =========================================================
ws=newsheet(S_COVER, TEAL)
widths(ws,{"A":4,"B":20,"C":40,"D":8,"E":12,"F":13,"G":10,"H":24})
title_row(ws,"🧭 網羅チェック（領域マップ）— やるべき/議論すべきの『基準』",8,TEAL)
sub_row(ws,"プロジェクトで必ず扱う領域のテンプレ。自社用に編集可。未着手の必須／久しく未議論＝抜けの兆候。",8)
CS=5; CE=40
ws.merge_cells("A3:B3"); cell(ws,3,1,"未着手の必須領域",bold=True,align="center",fillhex=C_DIV)
cell(ws,3,3,f'=COUNTIFS(D{CS}:D{CE},"必須",E{CS}:E{CE},"未着手")&" 件"',bold=True,align="center",fillhex=C_INPUT,color=RED)
ws.merge_cells("D3:E3"); cell(ws,3,4,"久しく未議論(必須>30日)",bold=True,align="center",fillhex=C_DIV)
ws.merge_cells("F3:G3"); cell(ws,3,6,f'=SUMPRODUCT((D{CS}:D{CE}="必須")*(F{CS}:F{CE}<>"")*((TODAY()-F{CS}:F{CE})>30)*(E{CS}:E{CE}<>"完了"))&" 件"',bold=True,align="center",fillhex=C_INPUT,color=AMBER)
hdr(ws,4,[("A4","#"),("B4","領域"),("C4","この領域で扱うこと（例）"),("D4","必須?"),("E4","状態"),("F4","最終議論日"),("G4","担当"),("H4","メモ")],TEAL)
areas=[("目的・スコープ","何をやる/やらない・ゴール"),("成果物・要件","アウトプットと要件定義"),("スケジュール","工程・マイルストン・締切"),("体制・役割","誰が何を（RACI）"),("予算・コスト","費用・リソース配分"),("リスク・課題","想定リスクと対応策"),("品質・受入基準","完成の定義・テスト観点"),("関係者・コミュ","ステークホルダー・連絡設計"),("依存・前提・制約","外部依存・前提条件"),("意思決定・承認","誰がどう決める・承認フロー"),("運用・保守・引継ぎ","リリース後・ハンドオフ"),("法務・コンプラ・セキュリティ","契約・規程・情報保護")]
for i in range(36):
    r=CS+i
    if i<len(areas):
        b,c=areas[i]; cell(ws,r,1,i+1,align="center",fillhex=C_DIV); cell(ws,r,2,b,bold=True); cell(ws,r,3,c,italic=True,color=C_EX)
        cell(ws,r,4,"必須",align="center",fillhex=C_INPUT); cell(ws,r,5,"未着手",align="center",fillhex=C_INPUT)
        cell(ws,r,6,"",align="center",fillhex=C_INPUT); cell(ws,r,7,"",align="center",fillhex=C_INPUT); cell(ws,r,8,"",fillhex=C_INPUT)
    else:
        cell(ws,r,1,i+1,align="center")
        inp(ws,r,2,""); inp(ws,r,3,""); cell(ws,r,4,"",align="center",fillhex=C_INPUT); cell(ws,r,5,"",align="center",fillhex=C_INPUT)
        cell(ws,r,6,"",align="center",fillhex=C_INPUT); cell(ws,r,7,"",align="center",fillhex=C_INPUT); inp(ws,r,8,"")
    ws[f"F{r}"].number_format="yyyy-mm-dd"
dv=DataValidation(type="list",formula1='"必須,任意"',allow_blank=True); ws.add_data_validation(dv); dv.add(f"D{CS}:D{CE}")
dv=DataValidation(type="list",formula1='"未着手,対応中,完了,対象外"',allow_blank=True); ws.add_data_validation(dv); dv.add(f"E{CS}:E{CE}")
crng=f"A{CS}:H{CE}"
ws.conditional_formatting.add(crng,FormulaRule(formula=[f'$E{CS}="完了"'],fill=fill(C_DONE)))
ws.conditional_formatting.add(crng,FormulaRule(formula=[f'AND($D{CS}="必須",$E{CS}="未着手")'],fill=fill(C_OVERDUE),font=Font(bold=True,color=RED)))
ws.conditional_formatting.add(crng,FormulaRule(formula=[f'$E{CS}="対象外"'],fill=fill("F1F5F9"),font=Font(color=C_EX)))
ws.conditional_formatting.add(f"F{CS}:F{CE}",FormulaRule(formula=[f'AND($D{CS}="必須",$F{CS}<>"",(TODAY()-$F{CS})>30,$E{CS}<>"完了")'],fill=fill(C_ISSUE),font=Font(bold=True,color=AMBER)))
ws.freeze_panes="A5"

# =========================================================
# 📖 ガイド
# =========================================================
ws=newsheet(S_GUIDE, TEAL)
widths(ws,{"A":18,"B":30,"C":40,"D":18})
title_row(ws,"📖 ガイド — 使い方・手法・困り事（迷ったらここ）",4,TEAL)
sub_row(ws,"『世界で一番やさしい会議の教科書』(榊巻亮/日経BP)。会議は小さな技で劇的に良くなる。まず『確認』から。",4)
r=3
section(ws,r,"■ 使い方（4タブ）",4,SLATE); r+=1
for a,b in [("▶ 会議","1回の会議。上から①準備→②進行→③確認→④振り返り。終わったら複製して次回へ"),
            ("🗂 課題・タスク台帳","決定/タスク/論点を全部貯める唯一の箱。上部に棚卸し(期限超過/停滞/担当なし)が自動"),
            ("🧭 網羅チェック","必ず扱う領域の基準マップ。未着手の必須・久しく未議論を自動であぶり出す"),
            ("📖 ガイド","この説明ページ")]:
    cell(ws,r,1,a,bold=True); ws.merge_cells(f"B{r}:D{r}"); cell(ws,r,2,b); ws.row_dimensions[r].height=28; r+=1
section(ws,r,"■ 抜け漏れを防ぐ（2レイヤー）",4,RED); r+=1
for a,b in [("会議内：論点網羅","▶会議の論点リスト＋カバレッジ＋MECE観点チェックで、その会議で扱うべき論点を漏らさない"),
            ("仕事全体：取りこぼし防止","全決定/タスクを🗂台帳へ＝1元管理。棚卸しで期限超過・停滞・担当なしを自動あぶり出し"),
            ("仕事全体：網羅性","🧭網羅チェックの領域マップ(基準)と現実を突き合わせ。未着手の必須・久しく未議論＝抜けの兆候"),
            ("運用のコツ","会議冒頭で🗂台帳の前回分を棚卸し→今日の論点は🧭網羅チェックと台帳から拾う")]:
    cell(ws,r,1,a,bold=True); ws.merge_cells(f"B{r}:D{r}"); cell(ws,r,2,b); ws.row_dimensions[r].height=30; r+=1
section(ws,r,"■ 凡例（色の意味）",4,SLATE); r+=1
for lab,hx,desc in [("意見",C_OPINION,"発言・アイデア"),("論点",C_ISSUE,"what to decide＝問い"),("決定",C_DECISION,"決まったこと（最重要）"),("完了/達成",C_DONE,"ToDo完了・終了条件達成・結論"),("要注意",C_OVERDUE,"期限超過・未達・未対応あり")]:
    cell(ws,r,1,lab,align="center",bold=True,fillhex=hx); ws.merge_cells(f"B{r}:D{r}"); cell(ws,r,2,desc); r+=1
section(ws,r,"■ 手法ガイド",4,TEAL); r+=1
hdr(ws,r,[(f"A{r}","手法"),(f"B{r}","要点（何を・なぜ）"),(f"C{r}","やり方"),(f"D{r}","場所")],TEAL); r+=1
guide=[
 ("①確認のファシリ","決定/ToDo/終了条件を声に出して確認。グダグダの特効薬。役職不要の第一歩","『決まったことは○○ですね』『誰が・いつまでに・何を?』","▶③"),
 ("②スクライブ","議論を可視化。見えると噛み合う。言った言わないが消える","意見/論点/決定を書き分け。記号・略字で速く","▶②"),
 ("③終了条件","終わりの状態を最初に定義。ゴール無しはダラダラの元","達成判定できる具体状態で。例『部下に説明できる状態』","▶①③"),
 ("④時間配分","議題ごとに分数を決め宣言。予実が改善のタネ","アジェンダに分数→開始時に宣言→予実記録","▶①④"),
 ("⑤Prep(4P)","Purpose/Process/People/Property を事前設計","Purpose=終了条件から先に。準備した分だけ短く濃く","▶①"),
 ("⑥隠れない","進め方を提案して確認。若手は隠れファシリで問い直す","『この順でいいですか?』『追いつかないので教えて』","▶②"),
 ("⑦発散と収束","出す→絞るを分ける。混ぜると良案が死ぬ","『今は出す時間』『ここから決める時間』と宣言","▶②"),
 ("⑧課題解決5階層","事象→問題→原因→施策→効果 で議論を揃える","今どの階層かを可視化、下から順に合意","▶②"),
 ("⑨沈黙の5心理","黙る理由は5つ。理由別に振る","ついていけない/言葉にならない/遠慮/同意/興味なし","▶②"),
 ("⑩抜け漏れ対策","論点を先に出し切り、カバレッジとMECEで網羅を担保","論点リスト＋状態管理＋観点チェック","▶①③"),
]
for a,b,c,d in guide:
    cell(ws,r,1,a,bold=True); cell(ws,r,2,b); cell(ws,r,3,c); cell(ws,r,4,d); ws.row_dimensions[r].height=44; r+=1
section(ws,r,"■ 困り事 → 対策",4,RED); r+=1
hdr(ws,r,[(f"A{r}","困り事"),(f"B{r}","効く手法"),(f"C{r}:D{r}","処方箋")],RED); r+=1
trouble=[
 ("必要な論点が抜ける","論点リスト/MECE","冒頭で論点を全部出し『これで全部?』と合意。観点チェックで抜けを潰す"),
 ("全部議論できたか不明","カバレッジ管理","各論点に状態。③の未対応カウントが0かで判別"),
 ("誰も発言しない","沈黙5心理/スクライブ","理由を見極め指名＋安全な問い。論点を書くと出やすい"),
 ("一部だけ話す","隠れない","『他の方の意見も聞きたい』＋指名。論点を可視化"),
 ("脱線する","終了条件/駐車場","終了条件と論点を指さす。脱線ネタは駐車場へ退避"),
 ("何も決まらない","確認のファシリ","最後に決定とToDo(誰/いつ/何)を声に出す"),
 ("言ってない問題","スクライブ","決定を全員の前でその場記録（②→③自動集約）"),
 ("噛み合わない","スクライブ/5階層","『何の問?』を明確化。5階層の今どこか確認"),
 ("絞れない","発散と収束","出す/決めるを分け、評価基準を先に合意"),
 ("時間内に終わらない","時間配分","配分を宣言。延長/収束/次回/非同期 から選ぶ"),
 ("偉い人が読むだけ","終了条件/隠れファシリ","冒頭で終了条件確認＋Q&A設計『確認です』で問い直す"),
 ("目的が曖昧","Prep(Purpose)","事前に主催者へ『この会議のゴールは?』"),
 ("人が多すぎ","Prep(People)","“いないと困る”人で絞る。他は議事録共有"),
 ("毎回同じ議論","ToDo管理/振り返り","冒頭で前回ToDo確認。TryをToDo一覧で追う"),
 ("案が出ない","発散と収束","事前に非同期収集も。『今は発散』と宣言"),
 ("責任者が曖昧","確認(ToDo)","ToDoに必ず『誰が』。1タスク1オーナー"),
]
for a,b,c in trouble:
    cell(ws,r,1,a,bold=True); cell(ws,r,2,b); ws.merge_cells(f"C{r}:D{r}"); cell(ws,r,3,c); ws.row_dimensions[r].height=30; r+=1
ws.freeze_panes="A3"

if "Sheet" in wb.sheetnames: del wb["Sheet"]
out="/home/neo/office-work/会議ファシリテーション_テンプレ.xlsx"
wb.save(out); print("SAVED",out); print("SHEETS",wb.sheetnames)
